import re
from openpyxl import Workbook,load_workbook
import openpyxl
import datetime
import tkinter as tk
from tkinter import filedialog
from tkinter import *
from tkinter import ttk
from tkinter.messagebox import showinfo

wb = Workbook()
we = wb.active
we.title = "Flow Data"

def openFile():
    rowDate=1
    rowTime=1
    rowFlow=1
    flowList=[]

    tf = filedialog.askopenfilename(
        initialdir="C:/Users/MainFrame/Desktop/", 
        title="Open Text file", 
        filetypes=(("Text Files", "*.txt"),)
        )
    path.insert(END, tf)
    tf = open(tf)
    textFile = tf.read()
    tf.close()
    flowFull = re.findall("FLOW: "+"\d+\.?\d*", textFile)  
    dates = re.findall("[0-9]+-[0-9]+-[0-9]+", textFile) 
    times = re.findall("\s[0-2][0-9]?[:]+[0-9]+[:]+[0-9]+", textFile)  

    for date in dates:  
        addedYear = "20" + date
        flowDate = datetime.datetime.strptime(addedYear, "%Y-%m-%d").strftime("%d.%m.%Y")
        we.cell(row=rowDate,column=1,value=flowDate)
        rowDate=rowDate+1

    for time in times: 
        we.cell(row=rowTime,column=2,value=time)
        rowTime=rowTime+1

    for flowFull in flowFull:
        flowNumber = re.findall("\d+\.?\d*", flowFull)  
        for flowNumber in flowNumber:
            we.cell(row=rowFlow,column=3,value=flowNumber.replace('.', ','))
            rowFlow=rowFlow+1
            flowList.append(float(flowNumber))

    avg=sum(flowList) / len(flowList)
    txtMin.delete("1.0","end")
    txtAvg.delete("1.0","end")
    txtMax.delete("1.0","end")
    txtMin.insert(END, min(flowList))
    txtAvg.insert(END, avg)
    txtMax.insert(END, max(flowList))
    wb.save("Flow Data.xlsx")
    wb.close()

ws = Tk()
ws.title("Flow Data Regex")
ws['bg']='#dd4124'
ws.geometry('300x200')
ws.resizable(False, False)

lblMin = Label( ws,text="Min Value: "  )
lblMin.grid(row=1,column=1)
txtMin = Text(ws, width=8, height=1)
txtMin.grid(row=1,column=2, padx=10, pady=10)
lblAvg = Label( ws,text="Avg Value: "  )
lblAvg.grid(row=2,column=1)
txtAvg = Text(ws, width=8, height=1)
txtAvg.grid(row=2,column=2, padx=10, pady=10)
lblMax = Label( ws,text="Max Value: "  )
lblMax.grid(row=3,column=1)
txtMax = Text(ws, width=8, height=1)
txtMax.grid(row=3,column=2,padx=10, pady=10)

path = Entry(ws)
path.grid(row=4,column=1,padx=10, pady=10)
Button(
    ws, 
    text="Open File", 
    command=openFile,
    width=15
    ).grid(row=4,column=2,padx=10, pady=10)
    
ws.mainloop()